import threading
import os
import math
from decimal import Decimal, InvalidOperation

# Check if we're in a production/lightweight environment
USE_LIGHTWEIGHT = os.environ.get('DJANGO_USE_LIGHTWEIGHT', 'false').lower() == 'true'

# Import dependencies based on environment
if not USE_LIGHTWEIGHT:
    try:
        import pandas as pd
        import numpy as np
        HAS_PANDAS = True
    except ImportError:
        HAS_PANDAS = False
        USE_LIGHTWEIGHT = True
else:
    HAS_PANDAS = False

# Lightweight utility functions (for when pandas/numpy is not available)
def is_nan_value(value):
    """Check if value is NaN (works without pandas/numpy)"""
    if value is None:
        return True
    if isinstance(value, str) and value.strip().lower() in ['nan', 'null', '']:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return False

def clean_decimal_value(value):
    """Clean and convert value to decimal - lightweight version"""
    try:
        if is_nan_value(value):
            return Decimal('0.00')
        str_value = str(value).strip()
        cleaned = str_value.replace(',', '').replace('%', '').replace('$', '')
        if not cleaned:
            return Decimal('0.00')
        return Decimal(cleaned).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal('0.00')

def clean_int_value(value):
    """Clean and convert value to integer - lightweight version"""
    try:
        if is_nan_value(value):
            return 0
        str_value = str(value).strip()
        cleaned = str_value.replace(',', '').replace('.0', '')
        if not cleaned:
            return 0
        return int(float(cleaned))
    except (ValueError, TypeError):
        return 0

def clean_string_value(value):
    """Enhanced to handle NaN values without pandas"""
    if is_nan_value(value):
        return ""
    try:
        str_value = str(value).strip()
        return str_value if str_value.lower() not in ['nan', 'null', 'none'] else ""
    except (TypeError, AttributeError):
        return ""

# Thread local storage for current tenant
_thread_local = threading.local()

def set_current_tenant(tenant):
    """Set the current tenant in thread local storage"""
    _thread_local.tenant = tenant

def get_current_tenant():
    """Get the current tenant from thread local storage"""
    return getattr(_thread_local, 'tenant', None)

def clear_current_tenant():
    """Clear the current tenant from thread local storage"""
    if hasattr(_thread_local, 'tenant'):
        delattr(_thread_local, 'tenant')

def generate_employee_id(name: str, tenant_id: int, department: str = None) -> str:
    """
    Generate employee ID using format: First three letters-Department first two letters-Tenant id
    Example: Siddhant Marketing Analysis tenant_id 025 -> SID-MA-025
    
    In case of collision with same name, add postfix A, B, C
    Example: SID-MA-025-A, SID-MA-025-B, SID-MA-025-C
    """
    from ..models import EmployeeProfile
    import uuid
    
    if not name or str(name).strip() in ['', '0', 'nan', 'NaN', '-']:
        return str(uuid.uuid4())[:8]  # Random ID for empty names
    
    # Extract first three letters from name (uppercase)
    name_clean = ''.join(char for char in name.strip().upper() if char.isalpha())
    name_prefix = name_clean[:3].ljust(3, 'X')  # Pad with X if less than 3 letters
    
    # Extract first two letters from department (uppercase)
    if department and str(department).strip():
        dept_clean = ''.join(char for char in str(department).strip().upper() if char.isalpha())
        dept_prefix = dept_clean[:2].ljust(2, 'X')  # Pad with X if less than 2 letters
    else:
        dept_prefix = 'XX'  # Default if no department
    
    # Format tenant ID with leading zeros (3 digits)
    tenant_str = str(tenant_id).zfill(3)
    
    # Generate base employee ID
    base_id = f"{name_prefix}-{dept_prefix}-{tenant_str}"
    
    # Check for collisions and add postfix if needed
    collision_suffixes = ['', '-A', '-B', '-C', '-D', '-E', '-F', '-G', '-H', '-I', '-J']
    
    for suffix in collision_suffixes:
        candidate_id = f"{base_id}{suffix}"
        
        # Check if this ID already exists in the database for this tenant
        if not EmployeeProfile.objects.filter(tenant_id=tenant_id, employee_id=candidate_id).exists():
            return candidate_id
    
def generate_employee_id_bulk_optimized(employees_data: list, tenant_id: int) -> dict:
    """
    ULTRA-FAST bulk employee ID generation for large datasets
    
    Process all employees in memory first, then generate unique IDs in batch
    This avoids N database queries during ID generation
    
    Args:
        employees_data: List of dicts with 'name', 'department' keys
        tenant_id: Tenant ID
    
    Returns:
        Dict mapping array index to generated employee_id
    """
    from ..models import EmployeeProfile
    import uuid
    from collections import defaultdict
    
    # Get all existing employee IDs for this tenant in one query
    existing_ids = set(
        EmployeeProfile.objects.filter(tenant_id=tenant_id)
        .values_list('employee_id', flat=True)
    )
    
    # Track generated IDs to avoid duplicates within this batch
    generated_ids = set()
    id_collision_counters = defaultdict(int)  # Track collision counts per base ID
    result_mapping = {}
    
    for index, emp_data in enumerate(employees_data):
        name = emp_data.get('name', '')
        department = emp_data.get('department', '')
        
        # Handle empty names
        if not name or str(name).strip() in ['', '0', 'nan', 'NaN', '-']:
            unique_id = str(uuid.uuid4())[:8]
            result_mapping[index] = unique_id
            generated_ids.add(unique_id)
            continue
        
        # Extract first three letters from name (uppercase)
        name_clean = ''.join(char for char in name.strip().upper() if char.isalpha())
        name_prefix = name_clean[:3].ljust(3, 'X')
        
        # Extract first two letters from department (uppercase)
        if department and str(department).strip():
            dept_clean = ''.join(char for char in str(department).strip().upper() if char.isalpha())
            dept_prefix = dept_clean[:2].ljust(2, 'X')
        else:
            dept_prefix = 'XX'
        
        # Format tenant ID with leading zeros (3 digits)
        tenant_str = str(tenant_id).zfill(3)
        
        # Generate base employee ID
        base_id = f"{name_prefix}-{dept_prefix}-{tenant_str}"
        
        # Check for collisions in existing DB + already generated IDs
        collision_suffixes = ['', '-A', '-B', '-C', '-D', '-E', '-F', '-G', '-H', '-I', '-J']
        
        candidate_id = None
        for suffix in collision_suffixes:
            test_id = f"{base_id}{suffix}"
            
            # Check if this ID exists in DB or already generated in this batch
            if test_id not in existing_ids and test_id not in generated_ids:
                candidate_id = test_id
                break
        
        # If all suffixes exhausted, use UUID fallback
        if not candidate_id:
            candidate_id = str(uuid.uuid4())[:8]
        
        result_mapping[index] = candidate_id
        generated_ids.add(candidate_id)
    
    return result_mapping

def validate_excel_columns(df_columns, required_columns):
    """
    Validate that all required columns are present in the Excel file
    """
    missing_columns = set(required_columns) - set(df_columns)
    if missing_columns:
        return False, f"Missing required columns: {', '.join(missing_columns)}"
    return True, "All required columns present"

def clean_decimal_value(value):
    """
    Clean and convert value to decimal - lightweight version
    """
    from decimal import Decimal
    try:
        # Handle NaN values first
        if not lightweight_notna(value):
            return Decimal('0.00')
            
        # Remove any commas and convert to string
        clean_value = str(value).replace(',', '').strip()
        
        # Handle empty string after cleaning
        if not clean_value or clean_value.lower() in ['nan', 'none', 'null']:
            return Decimal('0.00')
            
        return Decimal(clean_value)
    except (ValueError, TypeError, OverflowError):
        return Decimal('0.00')

def clean_int_value(value):
    """
    Clean and convert value to integer - lightweight version
    """
    try:
        # Handle NaN values first
        if not lightweight_notna(value):
            return 0
            
        # Remove any commas and convert to string
        clean_value = str(value).replace(',', '').strip()
        
        # Handle empty string after cleaning
        if not clean_value or clean_value.lower() in ['nan', 'none', 'null']:
            return 0
            
        return int(float(clean_value))
    except (ValueError, TypeError, OverflowError):
        return 0

def is_valid_name(name):
    """
    Check if a name is valid (not empty, not just '-', not '0', etc.)
    Lightweight version without pandas/numpy
    """
    # Handle NaN values
    if not lightweight_notna(name):
        return False
    
    if not name:
        return False
        
    name_str = str(name).strip()
    invalid_names = ['', '-', '0', 'nan', 'NaN', 'None', 'none', 'NULL', 'null']
    
    # Check if name is just one of the invalid values
    if name_str.lower() in [x.lower() for x in invalid_names]:
        return False
        
    # Check if name is only made up of special characters
    if all(c in '- _.,#@!$%^&*()' for c in name_str):
        return False
        
    return True 


# ==================== LIGHTWEIGHT PANDAS REPLACEMENTS ====================

def lightweight_notna(value):
    """Lightweight replacement for pd.notna()"""
    return not is_nan_value(value)

def lightweight_to_datetime(date_value):
    """Lightweight replacement for pd.to_datetime()"""
    from datetime import datetime
    
    if is_nan_value(date_value):
        return None
    
    try:
        if isinstance(date_value, datetime):
            return date_value.date()
        
        date_str = str(date_value).strip()
        if not date_str:
            return None
            
        # Common date formats
        formats = [
            '%Y-%m-%d',
            '%d/%m/%Y',
            '%m/%d/%Y',
            '%d-%m-%Y',
            '%Y/%m/%d',
            '%d/%m/%y',
            '%m/%d/%y',
            '%d-%m-%y',
            '%Y-%m-%d %H:%M:%S',
            '%d/%m/%Y %H:%M:%S',
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
                
        # Try to parse as Excel serial date
        try:
            excel_date = float(date_str)
            if excel_date > 25569:  # Excel epoch starts 1900-01-01
                from datetime import date, timedelta
                excel_epoch = date(1900, 1, 1)
                return excel_epoch + timedelta(days=int(excel_date) - 2)
        except (ValueError, TypeError):
            pass
            
        return None
    except Exception:
        return None

def lightweight_to_time(time_value):
    """Lightweight replacement for time parsing"""
    from datetime import time, datetime
    
    if is_nan_value(time_value):
        return None
    
    try:
        time_str = str(time_value).strip()
        if not time_str:
            return None
            
        # Common time formats
        formats = [
            '%H:%M:%S',
            '%H:%M',
            '%I:%M:%S %p',
            '%I:%M %p',
        ]
        
        for fmt in formats:
            try:
                dt = datetime.strptime(time_str, fmt)
                return dt.time()
            except ValueError:
                continue
                
        # Try to parse as Excel time serial
        try:
            time_float = float(time_str)
            if 0 <= time_float < 1:
                total_seconds = int(time_float * 24 * 60 * 60)
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                return time(hours, minutes, seconds)
        except (ValueError, TypeError):
            pass
            
        return None
    except Exception:
        return None

def excel_to_dict_list(file_obj, file_extension=None):
    """
    Convert Excel/CSV file to list of dictionaries (pandas-free).
    Handles various file formats and None values.
    """
    import openpyxl
    import csv
    import io
    
    try:
        # Determine file type
        if hasattr(file_obj, 'name'):
            filename = file_obj.name.lower()
        elif file_extension:
            filename = f"file.{file_extension.lower()}"
        else:
            filename = "file.xlsx"  # Default assumption
        
        if filename.endswith(('.xlsx', '.xls')):
            return _read_excel_lightweight(file_obj)
        elif filename.endswith('.csv'):
            return _read_csv_lightweight(file_obj)
        else:
            # Try Excel first, then CSV
            try:
                return _read_excel_lightweight(file_obj)
            except:
                file_obj.seek(0)  # Reset file pointer
                return _read_csv_lightweight(file_obj)
                
    except Exception as e:
        raise Exception(f"Error reading file: {str(e)}")

def _read_excel_lightweight(file_obj):
    """Read Excel file using openpyxl"""
    import openpyxl
    
    workbook = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    
    # Get headers from first row
    headers = []
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if first_row:
        headers = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(first_row)]
    
    # Convert rows to dictionaries
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row and any(cell is not None for cell in row):  # Skip empty rows
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = value
            data.append(row_dict)
    
    workbook.close()
    return data, headers

def _read_csv_lightweight(file_obj):
    """Read CSV file using csv module"""
    import csv
    import io
    
    # Handle different file object types
    if hasattr(file_obj, 'read'):
        content = file_obj.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8-sig')  # Handle BOM
        file_obj = io.StringIO(content)
    
    reader = csv.DictReader(file_obj)
    headers = reader.fieldnames or []
    data = list(reader)
    
    return data, headers

def filter_valid_rows(data, name_column='NAME'):
    """Filter out rows with invalid names (pandas-free)."""
    valid_rows = []
    for row in data:
        name = row.get(name_column, "")
        if is_valid_name(name):
            valid_rows.append(row)
    return valid_rows

def safe_float_conversion(value, default=0.0):
    """Safely convert value to float with fallback"""
    if is_nan_value(value):
        return default
    try:
        return float(str(value).strip().replace(',', ''))
    except (ValueError, TypeError):
        return default

def safe_int_conversion(value, default=0):
    """Safely convert value to int with fallback"""
    if is_nan_value(value):
        return default
    try:
        return int(float(str(value).strip().replace(',', '')))
    except (ValueError, TypeError):
        return default

def safe_str_conversion(value, default=''):
    """Safely convert value to string with fallback"""
    if is_nan_value(value):
        return default
    try:
        return str(value).strip()
    except (TypeError, AttributeError):
        return default

# Additional lightweight functions
def lightweight_to_numeric(value, errors='coerce'):
    """Lightweight replacement for pd.to_numeric()"""
    if is_nan_value(value):
        return 0.0 if errors == 'coerce' else None
    
    try:
        # Try converting to float first
        return float(str(value).strip().replace(',', ''))
    except (ValueError, TypeError):
        if errors == 'coerce':
            return 0.0
        elif errors == 'raise':
            raise ValueError(f"Unable to convert '{value}' to numeric")
        else:
            return value

def lightweight_fillna(data_list, column, fill_value):
    """Fill NaN values in a list of dictionaries for a specific column"""
    for row in data_list:
        if column in row and is_nan_value(row[column]):
            row[column] = fill_value
    return data_list

def create_test_excel_from_dict(data_list, filename, sheet_name='Sheet1'):
    """Create Excel file from list of dictionaries - lightweight alternative to DataFrame.to_excel()"""
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    if not data_list:
        wb.save(filename)
        return
    
    # Write headers
    headers = list(data_list[0].keys())
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Write data
    for row_idx, row_data in enumerate(data_list, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, '')
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(filename)

def dict_list_shape(data_list):
    """Get shape of list of dictionaries (rows, columns) - replaces df.shape"""
    if not data_list:
        return (0, 0)
    return (len(data_list), len(data_list[0]) if data_list[0] else 0)

def dict_list_columns(data_list):
    """Get column names from list of dictionaries - replaces df.columns"""
    if not data_list:
        return []
    return list(data_list[0].keys()) if data_list[0] else []

def dict_list_iterrows(data_list):
    """Iterate over list of dictionaries with index - replaces df.iterrows()"""
    for idx, row in enumerate(data_list):
        yield idx, row

# Create aliases for backward compatibility and ease of use
notna = lightweight_notna
to_datetime = lightweight_to_datetime
to_time = lightweight_to_time
to_numeric = lightweight_to_numeric
fillna = lightweight_fillna